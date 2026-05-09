#!/usr/bin/env python3
"""
MASANIELLO - EXTRACCIÓN Y TRADUCCIÓN A PYTHON

Sistema de gestión de riesgo Masaniello:
- Calcula montos dinámicos basados en W/L record
- Usa coeficiente multiplicador (L3) para incrementos
- Incorpora limitantes de riesgo (L4, L5)
- Modela ciclos de trading con objetivo de ganancias

PARÁMETROS CONFIGURABLES (células L1-L5):
  L1 = Objetivo de operaciones por ciclo (ej: 5)
  L2 = Objetivo de aciertos por ciclo (ej: 2)
  L3 = Coeficiente de incremento/multiplicador (ej: 1.5 o 2.0)
  L4 = Modo 0/1 (0=histórico, 1=sesión actual)
  L5 = Comisión o descuento % (ej: 2%)

FLUJO MASANIELLO:
1. Comienza con monto inicial (ej: $100)
2. Después de cada WIN: aumenta monto por (L3-1) * últimoMonto
3. Después de cada LOSS: se reduce el monto
4. Continúa hasta alcanzar L2 aciertos en L1 operaciones
5. Al completar ciclo: reset y vuelve al monto inicial

FÓRMULA CORE SIMPLIFICADA:
  Si es WIN:
    nuevo_monto = capital_actual + (capital_actual * (L3 - 1))
  Si es LOSS:
    nuevo_monto = capital_actual - (pérdida_calculada * comisión)
"""

import json
import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, asdict
from datetime import datetime

log = logging.getLogger("masaniello_engine")

# ──────────────────────────────────────────────────────────────────────────────
# DATOS Y ESTRUCTURAS
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class MasanielloConfig:
    """Configuración de parámetros Masaniello."""
    cycle_target_ops: int = 5          # L1: objetivo de operaciones/ciclo
    cycle_target_wins: int = 2         # L2: objetivo de ganancias/ciclo
    multiplier: float = 1.5            # L3: coeficiente multiplicador
    use_current_session: bool = False  # L4: 0=histórico, 1=sesión
    commission_pct: float = 2.0        # L5: comisión/descuento %
    
    initial_amount: float = 100.0      # Monto inicial de operación
    reference_balance: float = 100.0   # Banca base (F2/G2 en Excel)
    max_daily_loss: float = 500.0      # Límite de pérdida diaria
    payout_pct: float = 80.0           # Payout de ganancias en el broker
    excel_mirror_enabled: bool = False
    excel_mirror_path: str = ""
    excel_mirror_sheet: str = "Calcolatore"
    excel_mirror_column: str = "B"
    excel_mirror_start_row: int = 3

@dataclass
class TradeRecord:
    """Registro de una operación (W/L)."""
    cycle: int
    trade_num: int
    result: str  # "W" o "L"
    amount_risked: float
    payout_earned: float = 0.0
    timestamp: str = ""

@dataclass
class MasanielloState:
    """Estado actual del sistema Masaniello."""
    cycle_num: int = 1
    trades_in_cycle: int = 0
    wins_in_cycle: int = 0
    losses_in_cycle: int = 0
    current_capital: float = 100.0
    daily_loss_accumulated: float = 0.0
    total_pnl: float = 0.0
    bankroll: float = 100.0        # Columna F (banca)
    peak_bankroll: float = 100.0   # Columna G (máximo de banca)
    history: List[TradeRecord] = None
    
    def __post_init__(self):
        if self.history is None:
            self.history = []


# ──────────────────────────────────────────────────────────────────────────────
# MOTOR MASANIELLO
# ──────────────────────────────────────────────────────────────────────────────

class MasanielloEngine:
    """Motor de cálculo dinámico de montos basado en Masaniello."""
    
    def __init__(self, config: MasanielloConfig = None):
        self.config = config or MasanielloConfig()
        base = max(0.0, float(self.config.reference_balance))
        first_amount = max(0.0, float(self.config.initial_amount))
        self.state = MasanielloState(
            current_capital=first_amount,
            bankroll=base,
            peak_bankroll=base,
        )
        self._excel_next_row: Optional[int] = None
        self._excel_last_error: str = ""

    def _l3(self) -> float:
        return max(1.01, float(self.config.multiplier))

    def _l5(self) -> float:
        return max(0.0, float(self.config.commission_pct))

    def _excel_investment_amount(self, bankroll: float, losses: int, wins: int) -> float:
        """Replica la fórmula de inversión de la columna C del Excel Masaniello."""
        ops = max(1, int(self.config.cycle_target_ops))
        target = max(1, min(int(self.config.cycle_target_wins), ops))
        done = max(0, int(losses) + int(wins))
        l3 = self._l3()

        if wins >= target:
            return 0.0
        if losses >= (1 + ops - target):
            return 0.0

        values = self._excel_value_table(ops=ops, target=target, l3=l3)
        v_loss = values[min(done + 1, ops)][wins]
        v_win = values[min(done + 1, ops)][min(wins + 1, target)]
        denom = v_loss + (l3 - 1.0) * v_win
        if denom <= 0:
            return max(0.0, float(bankroll))

        fraction = 1.0 - ((l3 * v_win) / denom)
        return max(0.0, float(bankroll) * fraction)

    def _mirror_trade_result_to_excel(self, result: str) -> None:
        """Escribe W/L en Excel para comparación manual con la plantilla original."""
        if not self.config.excel_mirror_enabled:
            return

        excel_path = str(self.config.excel_mirror_path or "").strip()
        if not excel_path:
            return

        try:
            from openpyxl import load_workbook
        except Exception:
            if self._excel_last_error != "openpyxl_missing":
                self._excel_last_error = "openpyxl_missing"
                log.warning("Excel mirror desactivado: falta openpyxl.")
            return

        try:
            wb = load_workbook(excel_path)
            sheet_name = str(self.config.excel_mirror_sheet or "").strip()
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

            col = str(self.config.excel_mirror_column or "A").strip().upper()
            if not col.isalpha():
                col = "A"

            if self._excel_next_row is None:
                row = max(1, int(self.config.excel_mirror_start_row))
                while ws[f"{col}{row}"].value not in (None, ""):
                    row += 1
                self._excel_next_row = row

            cell_ref = f"{col}{self._excel_next_row}"
            ws[cell_ref] = "w" if str(result).upper() == "W" else "l"
            wb.save(excel_path)
            self._excel_next_row += 1
        except Exception as exc:
            err = str(exc)
            if err != self._excel_last_error:
                self._excel_last_error = err
                log.warning("No se pudo escribir W/L en Excel (%s): %s", excel_path, exc)

    @staticmethod
    def _excel_value_table(ops: int, target: int, l3: float) -> list[list[float]]:
        """Construye la tabla equivalente al bloque N:DJ del Excel Masaniello."""
        values: list[list[float]] = [
            [0.0 for _ in range(target + 2)]
            for __ in range(ops + 2)
        ]

        for done in range(ops, -1, -1):
            rem = ops - done
            for wins in range(target, -1, -1):
                if wins >= target:
                    values[done][wins] = 1.0
                    continue

                required = target - wins
                if required == rem:
                    values[done][wins] = l3 ** rem
                    continue

                if rem == 0:
                    values[done][wins] = 0.0
                    continue

                v_loss = values[done + 1][wins]
                v_win = values[done + 1][wins + 1]
                denom = v_loss + (l3 - 1.0) * v_win
                values[done][wins] = 0.0 if denom <= 0 else (l3 * v_loss * v_win) / denom

        return values

    def calculate_initial_amount_from_balance(self, balance: float) -> float:
        """Calcula primera inversión con la misma lógica base del Excel Masaniello."""
        return self._excel_investment_amount(bankroll=float(balance), losses=0, wins=0)
    
    def calculate_next_amount(self, previous_amount: float, is_win: bool) -> float:
        """
        Calcula el monto de la siguiente operación.
        
        FÓRMULA MASANIELLO:
          Si WIN: nuevo_monto = monto_actual + (monto_actual * (L3 - 1))
                               = monto_actual * L3
          Si LOSS: nuevo_monto = monto_actual - descuento
        
        Args:
            previous_amount: Monto de la operación anterior
            is_win: True si fue ganancia, False si fue pérdida
            
        Returns:
            Monto calculado para la siguiente operación
        """
        # Conservado por compatibilidad externa; el flujo real usa process_trade().
        losses = int(self.state.losses_in_cycle) + (0 if is_win else 1)
        wins = int(self.state.wins_in_cycle) + (1 if is_win else 0)
        amount = self._excel_investment_amount(
            bankroll=float(self.state.bankroll),
            losses=losses,
            wins=wins,
        )
        return round(amount, 2)
    
    def process_trade(self, result: str, amount_risked: float, 
                     payout_earned: float = 0.0) -> Dict:
        """
        Procesa una operación y actualiza el estado.
        
        Args:
            result: "W" para win, "L" para loss
            amount_risked: Dinero en riesgo en esta operación
            payout_earned: Ganancia neta si es WIN
            
        Returns:
            Dict con información de la operación procesada
        """
        is_win = result.upper() == "W"
        
        # Actualizar contadores
        self.state.trades_in_cycle += 1
        if is_win:
            self.state.wins_in_cycle += 1
            self.state.total_pnl += payout_earned
        else:
            self.state.losses_in_cycle += 1
            loss_amount = amount_risked * (1 - self.config.payout_pct / 100)
            self.state.daily_loss_accumulated += loss_amount
            self.state.total_pnl -= loss_amount
        
        # Registrar en histórico
        trade_rec = TradeRecord(
            cycle=self.state.cycle_num,
            trade_num=self.state.trades_in_cycle,
            result=result,
            amount_risked=amount_risked,
            payout_earned=payout_earned if is_win else -amount_risked,
            timestamp=datetime.now().isoformat()
        )
        self.state.history.append(trade_rec)
        self._mirror_trade_result_to_excel(result)
        
        # Replica Excel (columnas D/F/G): actualizar banca con el resultado real.
        l3 = self._l3()
        l5 = self._l5()
        b = 1 if is_win else 0
        f_prev = float(self.state.bankroll)
        g_prev = float(self.state.peak_bankroll)
        c_used = max(0.0, float(amount_risked))
        d_profit = c_used * (l3 - 1.0) * b

        if b == 0:
            f_new = f_prev - c_used
        else:
            gap_to_peak = g_prev - f_prev
            if (f_prev + d_profit - g_prev) >= 0:
                f_new = f_prev + gap_to_peak + ((d_profit - gap_to_peak) * (l5 / 100.0))
            else:
                f_new = f_prev + d_profit
        g_new = max(f_new, g_prev)
        self.state.bankroll = float(f_new)
        self.state.peak_bankroll = float(g_new)

        # Calcular próximo monto con la misma ecuación de inversión del Excel (columna C).
        next_amount = self._excel_investment_amount(
            bankroll=float(self.state.bankroll),
            losses=int(self.state.losses_in_cycle),
            wins=int(self.state.wins_in_cycle),
        )

        # Mantener capital actual sincronizado con el monto que debe usar la siguiente entrada.
        self.state.current_capital = float(next_amount)

        # Capturar secuencia y métricas del ciclo antes de un posible reset.
        cycle_num_before_reset = int(self.state.cycle_num)
        trades_before_reset = int(self.state.trades_in_cycle)
        wins_before_reset = int(self.state.wins_in_cycle)
        losses_before_reset = int(self.state.losses_in_cycle)
        seq_before_reset = "".join(
            str(t.result).upper()
            for t in self.state.history
            if int(t.cycle) == cycle_num_before_reset
        )
        
        # Verificar si ciclo está completo
        # Regla 2-de-5 matemáticamente consistente:
        # para 5/2 el ciclo aún puede recuperarse hasta 3 pérdidas, y corta en 4 pérdidas.
        max_losses_cycle_cutoff = max(1, 1 + int(self.config.cycle_target_ops) - int(self.config.cycle_target_wins))
        cycle_complete = (
            self.state.trades_in_cycle >= self.config.cycle_target_ops or
            self.state.wins_in_cycle >= self.config.cycle_target_wins or
            self.state.losses_in_cycle >= max_losses_cycle_cutoff or
            next_amount <= 0.0
        )

        close_reason = ""
        if cycle_complete:
            if self.state.wins_in_cycle >= self.config.cycle_target_wins:
                close_reason = "target_wins_reached"
            elif self.state.losses_in_cycle >= max_losses_cycle_cutoff:
                close_reason = "loss_cutoff_reached"
            elif next_amount <= 0.0:
                close_reason = "non_operable_next_amount"
            else:
                close_reason = "max_ops_reached"
        
        if cycle_complete:
            self._complete_cycle()
        
        return {
            "trade_num": trades_before_reset,
            "cycle": cycle_num_before_reset,
            "result": result,
            "amount_risked": amount_risked,
            "next_amount": next_amount,
            "wins": wins_before_reset,
            "losses": losses_before_reset,
            "total_pnl": self.state.total_pnl,
            "cycle_complete": cycle_complete,
            "sequence": seq_before_reset,
            "close_reason": close_reason,
        }
    
    def _complete_cycle(self):
        """Finaliza un ciclo y prepara el siguiente."""
        # Resumen del ciclo completado
        win_rate = self.state.wins_in_cycle / max(1, self.state.trades_in_cycle) * 100
        
        # Reset para nuevo ciclo
        self.state.cycle_num += 1
        self.state.trades_in_cycle = 0
        self.state.wins_in_cycle = 0
        self.state.losses_in_cycle = 0
        # Definición operativa solicitada: cada ciclo reinicia con banca base fija.
        self.state.bankroll = float(self.config.reference_balance)
        self.state.peak_bankroll = float(self.config.reference_balance)
        self.state.current_capital = self._excel_investment_amount(
            bankroll=float(self.config.reference_balance),
            losses=0,
            wins=0,
        )
        
        print(f"\n✓ Ciclo completado | WR: {win_rate:.0f}% | PnL: ${self.state.total_pnl:.2f}")
        print(f"→ Iniciando ciclo #{self.state.cycle_num}...\n")
    
    def get_status(self) -> Dict:
        """Retorna estado actual del motor."""
        return {
            "cycle": self.state.cycle_num,
            "trades_this_cycle": self.state.trades_in_cycle,
            "wins_this_cycle": self.state.wins_in_cycle,
            "losses_this_cycle": self.state.losses_in_cycle,
            "win_rate_pct": (self.state.wins_in_cycle / max(1, self.state.trades_in_cycle) * 100) if self.state.trades_in_cycle > 0 else 0,
            "total_pnl": self.state.total_pnl,
            "daily_loss": self.state.daily_loss_accumulated,
        }
    
    def export_history(self, filepath: str = "masaniello_history.json"):
        """Exporta el histórico de operaciones."""
        data = {
            "config": asdict(self.config),
            "state": {
                "cycle": self.state.cycle_num,
                "total_pnl": self.state.total_pnl,
                "trades_count": len(self.state.history),
            },
            "trades": [asdict(t) for t in self.state.history]
        }
        
        with open(filepath, "w") as f:
            json.dump(data, f, indent=2, default=str)
        
        print(f"✓ Histórico exportado a: {filepath}")


# ──────────────────────────────────────────────────────────────────────────────
# EJEMPLO DE USO
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("═" * 80)
    print("🔢 MOTOR MASANIELLO - DEMOSTRACIÓN")
    print("═" * 80)
    
    # Crear configuración
    config = MasanielloConfig(
        cycle_target_ops=5,      # 5 operaciones por ciclo
        cycle_target_wins=2,     # Objetivo: 2 ganancias
        multiplier=1.5,          # Incremento: 50%
        commission_pct=2.0,      # Comisión: 2%
        initial_amount=100.0,    # Monto inicial: $100
    )
    
    print("\n📋 CONFIGURACIÓN:")
    print(f"  Ciclo objetivo: {config.cycle_target_ops} ops, {config.cycle_target_wins} wins")
    print(f"  Multiplicador: {config.multiplier}x")
    print(f"  Monto inicial: ${config.initial_amount}")
    print(f"  Comisión: {config.commission_pct}%")
    
    # Crear motor
    engine = MasanielloEngine(config)
    
    # Simular 10 operaciones
    print("\n📊 SIMULACIÓN DE OPERACIONES:")
    print("-" * 80)
    print("Op | Res | Monto Riesgo | Siguiente | W | L | P&L")
    print("-" * 80)
    
    trades_simulation = [
        ("W", 100.0, 80.0),    # Op 1: WIN
        ("W", 150.0, 120.0),   # Op 2: WIN → Ciclo completo (2/2 wins)
        ("L", 100.0, 0.0),     # Op 3 (Ciclo 2): LOSS
        ("W", 100.0, 80.0),    # Op 4: WIN
        ("L", 150.0, 0.0),     # Op 5: LOSS
        ("L", 98.0, 0.0),      # Op 6: LOSS
        ("W", 96.0, 77.0),     # Op 7: WIN
        ("W", 144.0, 115.0),   # Op 8: WIN → Ciclo completo
        ("W", 100.0, 80.0),    # Op 9 (Ciclo 3): WIN
        ("W", 150.0, 120.0),   # Op 10: WIN → Ciclo completo
    ]
    
    for op_num, (result, amount, payout) in enumerate(trades_simulation, 1):
        result_obj = engine.process_trade(result, amount, payout if result == "W" else 0)
        
        w_count = result_obj["wins"]
        l_count = result_obj["losses"]
        pnl = result_obj["total_pnl"]
        next_amt = result_obj["next_amount"]
        
        marker = " ✓" if result_obj["cycle_complete"] else ""
        print(f"{op_num:2d} | {result:^3} | ${amount:8.2f} | ${next_amt:8.2f} | {w_count} | {l_count} | ${pnl:7.2f}{marker}")
    
    # Estado final
    print("-" * 80)
    status = engine.get_status()
    print(f"\n📈 ESTADO FINAL:")
    print(f"  Ciclo actual: #{status['cycle']}")
    print(f"  Operaciones en ciclo: {status['trades_this_cycle']}")
    print(f"  Wins/Losses: {status['wins_this_cycle']}/{status['losses_this_cycle']}")
    print(f"  Win Rate: {status['win_rate_pct']:.1f}%")
    print(f"  Total P&L: ${status['total_pnl']:.2f}")
    
    # Exportar
    export_path = "data/masaniello_demo.json"
    Path(export_path).parent.mkdir(parents=True, exist_ok=True)
    engine.export_history(export_path)
    
    print("\n" + "═" * 80)
    print("\n✓ Lógica Masaniello lista para integración en martingale_calculator.py")
    print("  Ver documentación en: MASANIELLO_TRANSLATION.md")
